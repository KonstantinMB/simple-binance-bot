from binance.client import Client, BaseClient
from pandas import DataFrame
import pandas as pd # Run `pip install pyarrow` if deprecation notice occurs
import time
import openpyxl
import datetime
import numpy as np
import matplotlib.pyplot as plt
import mplfinance as mpf

# To load the API_KEYs & API_SECRETs, we load them from a .env file
from dotenv import load_dotenv
import os
load_dotenv() 

# Configs for DEMO Account
test_api_key=os.getenv('TEST_API_KEY')
test_api_secret=os.getenv('TEST_API_SECRET')

# Configs for REAL
api_key = os.getenv('API_KEY')
api_secret = os.getenv('API_SECRET')

# Configs for logging & saving data
config_path = str(os.getenv('CONFIG_PATH'))
chart_config_path = config_path + "Chart.png"
intra_day_excel = config_path + "IntraDayData.xlsx"

# Trade Configurations
symbol_pair = 'BTCUSDT'
candle_timeframe = "1m"
buy_amount = 100
demo = True
moving_average_span = 26
order_id = 0
stop_loss_percent = 0.02
take_profit_percent = 0.1
HHV = 10
LLV = 10

# Do you want to add stop_loss/take_profit prices once the order is FILLED 👇
stop_trigger_toggle = True

# We support only MARKET (for MARKET_ORDER) or LIMIT (for LIMIT_MARKET_ORDER) orders 👇
order_type = 'LIMIT'

price_precision, quantity_precision, order_quantity = 0, 0, 0


def boundaryRemaining(tf):
        # "1m", "3m", "5m", "15m", "30m", "1h", "2h", "4h", "6h", "8h", "12h", "1d", "3d", "1w", "1M"
        #t = datetime.datetime.now()
        t = datetime.datetime.utcnow()
        if tf == "1m":
            next_t = (t+datetime.timedelta(minutes=1)).replace(second=0,microsecond=0)
        elif tf == "3m":
            next_t = (t+datetime.timedelta(minutes=3-t.minute%3)).replace(second=0, microsecond=0)
        elif tf == "5m":
            next_t = (t+datetime.timedelta(minutes=5-t.minute%5)).replace(second=0, microsecond=0)
        elif tf == "15m":
            next_t = (t+datetime.timedelta(minutes=15-t.minute%15)).replace(second=0, microsecond=0)
        elif tf == "30m":
            next_t = (t+datetime.timedelta(minutes=30-t.minute%30)).replace(second=0, microsecond=0)
        elif tf == "1h":
            next_t = (t+datetime.timedelta(hours=1)).replace(minute=0,second=0,microsecond=0)
        elif tf == "2h":
            next_t = (t+datetime.timedelta(hours=2-t.hour%2)).replace(minute=0, second=0, microsecond=0)
        elif tf == "4h":
            next_t = (t+datetime.timedelta(hours=4-t.hour%4)).replace(minute=0, second=0, microsecond=0)
        elif tf == "6h":
            next_t = (t+datetime.timedelta(hours=6-t.hour%6)).replace(minute=0, second=0, microsecond=0)
        elif tf == "8h":
            next_t = (t+datetime.timedelta(hours=8-t.hour%8)).replace(minute=0, second=0, microsecond=0)
        elif tf == "12h":
            next_t = (t+datetime.timedelta(hours=12-t.hour%12)).replace(minute=0, second=0, microsecond=0)
        elif tf == "1d":
            next_t = (t+datetime.timedelta(hours=24)).replace(hour=0,minute=0, second=0, microsecond=0)
        elif tf == "3d":
            day_pivot = datetime.datetime(2017, 8, 17, 0, 0)
            next_t = (t+datetime.timedelta(days=3-(t-day_pivot).days%3)).replace(hour=0,minute=0, second=0, microsecond=0)
        elif tf == "1w":  # Monday 0am
            next_t = (t+datetime.timedelta(days=7-t.weekday())).replace(hour=0,minute=0,second=0,microsecond=0)
        elif tf == "1M":  # 0am
            if t.month==12:
                next_t = datetime.datetime(t.year+1,1,1,0,0,0)
            else:
                next_t = datetime.datetime(t.year, t.month+1, 1, 0, 0, 0)
        remaining = next_t-t
        return remaining

def cur_time():
    s = "[" + time.strftime("%d%b%Y", time.localtime()) + "]"
    s = s+"[" + time.strftime("%H:%M:%S", time.localtime()) + "]"
    return s.upper()

def today(length = 6):
    if length==8:
        return time.strftime("%Y%m%d", time.localtime())
    elif length==6:
        return time.strftime("%y%m%d", time.localtime())
    elif length==4:
        return time.strftime("%m%d", time.localtime())
    
def std_log(s):
    global config_path
    print(cur_time() + s)
    fout = open(config_path + "log%s.txt"%(today()), "a")
    fout.writelines(cur_time()+s+"\n")
    fout.close()

def update_excel_with_new_data(filename, symbol, new_data_df: DataFrame):
    # Ensure new_data_df is a DataFrame and not empty
    if new_data_df.empty:
        print(f"No data provided for {symbol}.")
        return

    if os.path.isfile(filename):
        # Load the existing workbook
        book = openpyxl.load_workbook(filename)
        
        # Check if it's the only sheet; if so, add a temporary one
        if symbol in book.sheetnames and len(book.sheetnames) == 1:
            book.create_sheet(title="TempSheetForDeletion")
        
        # Remove the sheet for the symbol if it exists
        if symbol in book.sheetnames:
            del book[symbol]

        # Save the changes to the workbook
        book.save(filename)
        book.close()
    
    # Write (or overwrite) the data to the workbook
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a' if os.path.exists(filename) else 'w') as writer:
        new_data_df.to_excel(writer, sheet_name=symbol, index=False)

        # If a temporary sheet was added, remove it now
        if "TempSheetForDeletion" in writer.book.sheetnames:
            del writer.book["TempSheetForDeletion"]

def fetch_and_store_data(client: Client, symbol, timeframe):
    # Fetch the klines (OHLC data) from Binance Futures
    klines = client.futures_klines(symbol=symbol, interval=timeframe)

    # Process each kline entry to extract and convert OHLC, timestamp, and volume
    ohlc_data = [
        [
            datetime.datetime.utcfromtimestamp(entry[0] / 1000),  # Timestamp used for indexing
            datetime.datetime.utcfromtimestamp(entry[0] / 1000).strftime('%Y-%m-%d %H:%M:%S'),  # Timestamp
            float(entry[1]),  # Open price
            float(entry[2]),  # High price
            float(entry[3]),  # Low price
            float(entry[4]),  # Close price
            float(entry[5])   # Volume
        ]
        for entry in klines
    ]

    # Create DataFrame with OHLC data, timestamp, and volume
    df = pd.DataFrame(ohlc_data, columns=['index_timestamp', 'timestamp', 'open', 'high', 'low', 'close', 'volume'])
    
    # Set timestmap as index for later use:
    df.set_index('index_timestamp', inplace=True)

    return df

def get_precision_for_symbol(client: Client, symbol):

    exchange_info = client.futures_exchange_info()

    for symbol_info in exchange_info['symbols']:
        if symbol_info['symbol'] == symbol:
            price_precision = None
            quantity_precision = None
            for filter in symbol_info['filters']:
                if filter['filterType'] == 'PRICE_FILTER':
                    price_precision = filter['tickSize'].find('1') - 1
                elif filter['filterType'] == 'LOT_SIZE':
                    quantity_precision = filter['stepSize'].find('1') - 1
            return price_precision, quantity_precision
        
    return None, None  # In case the symbol is not found

def place_new_futures_market_order(client: Client, symbol_pair, buy_amount,
                                   positionSide='LONG', side=BaseClient.SIDE_BUY):
    
    global quantity_precision, order_quantity
    _, quantity_precision = get_precision_for_symbol(client, symbol_pair)

    latest_price_info = client.futures_symbol_ticker(symbol=symbol_pair)
    latest_price = float(latest_price_info['price'])
    order_quantity = round(buy_amount / latest_price, quantity_precision)

    try:
        order_info = client.futures_create_order(
            symbol=symbol_pair,
            side=side,
            positionSide=positionSide,
            type=BaseClient.ORDER_TYPE_MARKET,  # Change to market order type
            quantity=order_quantity,
            recvWindow=10000
        )

        std_log(f"[{symbol_pair}] MARKET Buy Order Executed. Order Info: {order_info}")
        return order_info

    except Exception as e:
        std_log(f"[{symbol_pair}] An error occurred while making a market order. Error Info: {e}")
        return None

def place_new_futures_limit_order(client: Client, symbol_pair, buy_amount, limit_price,
                        positionSide = 'LONG',
                        side = BaseClient.SIDE_BUY, 
                        type = BaseClient.FUTURE_ORDER_TYPE_LIMIT):

    global price_precision, quantity_precision, order_quantity
    price_precision, quantity_precision = get_precision_for_symbol(client, symbol_pair)
    price = round(limit_price, price_precision)
    order_quantity = round((buy_amount / limit_price), quantity_precision)

    try:

        order_info = client.futures_create_order(
            symbol=symbol_pair, 
            side=side,
            positionSide=positionSide,
            type=type,
            price=price,
            quantity=order_quantity,
            recvWindow= 10000000,
            timeInForce='GTC'
        )

        std_log("[%s] LIMIT Buy Order Executed. Order Info: %s" %(symbol_pair, str(order_info)))
        return order_info
    
    except Exception as e:
        std_log("[%s] An error occured while making an order. Error Info: [%s]" %(symbol_pair, e))
        return None

def check_order_status(client: Client, symbol_pair, order_id):

    try:
        order_status = client.futures_get_order(symbol=symbol_pair, orderId=order_id)
        std_log("[%s] Order ID [%s] with status: [%s]" %(symbol_pair, order_id, order_status['status']))
        return order_status
    
    except Exception as e:
        std_log(f"[{symbol_pair}] Error checking order status. Error Info: {e}")
        return None
    
def get_crypto_data(client: Client, symbol_pair, timeframe):

    chart_df = fetch_and_store_data(client, symbol_pair, timeframe)

    if chart_df.index[-1] > datetime.datetime.utcnow() - buy_timedelta[symbol_pair]/2:
                chart_df = chart_df.iloc[:-1] 
    
    chart_df['ema'] = round(chart_df['close'].ewm(span=moving_average_span, adjust=False).mean(), 4)
    
    # Bollinger Bands
    chart_df['20 Day MA'] = chart_df['close'].rolling(window=20).mean()
    chart_df['20 Day STD'] = chart_df['close'].rolling(window=20).std()
    chart_df['Upper Band'] = chart_df['20 Day MA'] + (chart_df['20 Day STD'] * 2)
    chart_df['Lower Band'] = chart_df['20 Day MA'] - (chart_df['20 Day STD'] * 2)

    # Simple Moving Averages
    chart_df['SMA_10'] = chart_df['close'].rolling(window=10).mean()
    chart_df['SMA_50'] = chart_df['close'].rolling(window=50).mean()
    chart_df['SMA_200'] = chart_df['close'].rolling(window=200).mean()

    # Exponential Moving Averages
    chart_df['EMA_12'] = chart_df['close'].ewm(span=12, adjust=False).mean()
    chart_df['EMA_26'] = chart_df['close'].ewm(span=26, adjust=False).mean()

    # Moving Average Convergence Divergence
    chart_df['MACD_line'] = chart_df['EMA_12'] - chart_df['EMA_26']
    chart_df['Signal_line'] = chart_df['MACD_line'].ewm(span=9, adjust=False).mean()
    chart_df['MACD_Hist'] = chart_df['MACD_line'] - chart_df['Signal_line']

    # Calculate ATR
    chart_df['TR'] = np.maximum(chart_df['high'] - chart_df['low'],
                                np.maximum(abs(chart_df['high'] - chart_df['close'].shift()),
                                           abs(chart_df['low'] - chart_df['close'].shift())))
    chart_df['ATR_20'] = chart_df['TR'].rolling(window=20).mean()

    # Calculate Highest High and Lowest Low
    chart_df['Highest_High_15'] = chart_df['high'].rolling(window=HHV).max()
    chart_df['Lowest_Low_10'] = chart_df['low'].rolling(window=LLV).min()

    update_excel_with_new_data(intra_day_excel, symbol_pair, chart_df)

    # We are making a visual chart for better trade visibility (Chart.png)
    plot_crypto_chart_with_indicators(chart_df, chart_config_path)

    return chart_df, chart_df['ema'].iloc[-1]

def plot_crypto_chart_with_indicators(chart_df: DataFrame, file_path: str, symbol_pair: str):
    
    # Ensure the DataFrame's index is a DatetimeIndex
    chart_df.index = pd.to_datetime(chart_df.index)
    
    # Define additional plots
    apds = [
        mpf.make_addplot(chart_df['ema'], color='blue'),
        mpf.make_addplot(chart_df['20 Day MA'], color='red'),
        mpf.make_addplot(chart_df['Upper Band'], color='green'),
        mpf.make_addplot(chart_df['Lower Band'], color='green'),
        mpf.make_addplot(chart_df['SMA_10'], color='orange'),
        mpf.make_addplot(chart_df['SMA_50'], color='purple'),
        mpf.make_addplot(chart_df['SMA_200'], color='black'),
        mpf.make_addplot(chart_df['MACD_line'], panel=1, color='fuchsia'),
        mpf.make_addplot(chart_df['Signal_line'], panel=1, color='b'),
        mpf.make_addplot(chart_df['MACD_Hist'], panel=1, type='bar', color='dimgray', alpha=0.3),
    ]

    # Define the market colors and style
    mc = mpf.make_marketcolors(up='green', down='red', inherit=True)
    s = mpf.make_mpf_style(marketcolors=mc)
    
    # Plotting
    mpf.plot(chart_df, type='candle', style=s, addplot=apds, volume=True, figratio=(12, 8), 
             title="\n" + symbol_pair + " Chart with Indicators", savefig=file_path)

def update_order_id(order_info):
    
    global order_id
    if order_info is not None:
        order_id = order_info['orderId']

    return order_id
    
def fetch_last_close_price(chart_df: DataFrame):

    return chart_df['close'].iloc[-1]

def modify_order(client: Client, symbol_pair, order_id, buy_amount, new_price):

    global order_type

    try:
        # First, cancel the original order
        cancel_response = client.futures_cancel_order(symbol=symbol_pair, orderId=order_id)
        std_log(f"[{symbol_pair}] Original order canceled. Info: {cancel_response}")

        # Then, place a new order with the updated price
        new_order_response = place_futures_order(client, symbol_pair, buy_amount, new_price, order_type)
        return new_order_response
    
    except Exception as e:
        std_log(f"[{symbol_pair}] Error modifying order. Error Info: {e}")
        return None
    
def place_stop_triggers(client: Client, symbol, last_close_price):

    global price_precision, quantity_precision, order_quantity
    take_profit_price = round(last_close_price * (1 + take_profit_percent), price_precision)
    stop_loss_price = round(last_close_price * (1 - stop_loss_percent), price_precision)

    take_profit_market_order= client.futures_create_order(
        symbol=symbol,
        side=BaseClient.SIDE_SELL,
        positionSide='LONG',
        type=BaseClient.FUTURE_ORDER_TYPE_TAKE_PROFIT_MARKET,
        timeInForce=BaseClient.TIME_IN_FORCE_GTC,
        stopPrice=take_profit_price,
        closePosition=True
        )
    
    std_log("[%s] TAKE_PROFIT_MARKET Buy Order Executed. Order Info: %s" %(symbol, str(take_profit_market_order)))

    stop_market_order = client.futures_create_order(
        symbol=symbol,
        side=BaseClient.SIDE_SELL,
        positionSide='LONG',
        type=BaseClient.FUTURE_ORDER_TYPE_STOP_MARKET,
        quantity=order_quantity,
        stopPrice=stop_loss_price,
        closePosition=True
    )

    std_log("[%s] STOP_MARKET Buy Order Executed. Order Info: %s" %(symbol, str(stop_market_order)))

def place_futures_order(client: Client, symbol_pair, buy_amount, initialEMA, order_type):

    new_order_info = ''

    if order_type == 'LIMIT':
        new_order_info = place_new_futures_limit_order(client, symbol_pair, buy_amount, initialEMA)

    elif order_type == 'MARKET': 
        new_order_info = place_new_futures_market_order(client, symbol_pair, buy_amount)

    else:
        std_log(f"[{symbol_pair}] Invalid ORDER TYPE {order_type}. Please use LIMIT or MARKET orders only!")

    update_order_id(new_order_info)

    return new_order_info

if __name__=="__main__":
    
    # Time Configurations For Trade Set-Up 👇 
    tdelta_conv = {"1m": datetime.timedelta(minutes=1), "3m": datetime.timedelta(minutes=3),
                "5m": datetime.timedelta(minutes=5), "15m": datetime.timedelta(minutes=15),
                "30m": datetime.timedelta(minutes=30), "1h": datetime.timedelta(hours=1),
                "2h": datetime.timedelta(hours=2), "4h": datetime.timedelta(hours=4),
                "6h": datetime.timedelta(hours=6), "8h": datetime.timedelta(hours=8),
                "12h": datetime.timedelta(hours=12), "1d": datetime.timedelta(days=1),
                "3d": datetime.timedelta(days=3), "1w": datetime.timedelta(days=7)}
    buy_timedelta = {}
    old_remain_long_buy = {}   
    buy_timedelta[symbol_pair] = tdelta_conv[candle_timeframe]  
    old_remain_long_buy[symbol_pair] = datetime.timedelta(days=365)

    # Initializing the Binance API client based on the `demo` param
    client = Client()
    if demo:
        client = Client(api_key=test_api_key, api_secret=test_api_secret, testnet=demo)
    else: 
        client = Client(api_key=api_key, api_secret=api_secret, testnet=demo)

    # Retrieve OHLC & Other Indicators Data
    chart_df, initialEMA = get_crypto_data(client, symbol_pair, candle_timeframe)
    last_close_price = fetch_last_close_price(chart_df)
    
    # Place First Order
    order_info = place_futures_order(client, symbol_pair, buy_amount, initialEMA, order_type)
    status = check_order_status(client, symbol_pair, order_id)

    while True:

        showed_remain = datetime.timedelta(days=365)
        remain = boundaryRemaining(candle_timeframe)   # Remain time to buy candle closing
        showed_remain = min(remain, showed_remain)
        
        if old_remain_long_buy[symbol_pair] < remain: # Get into new candle

            status = check_order_status(client, symbol_pair, order_id)

            if status['status'] != 'FILLED':

                # Retrieving latest OHCL & EMA data:
                chart_df, ema = get_crypto_data(client, symbol_pair, candle_timeframe)
                last_close_price = fetch_last_close_price(chart_df)

                std_log("[%s] New Bar: %s" %(symbol_pair, chart_df.iloc[-1]))

                new_order_info = modify_order(client=client, symbol_pair=symbol_pair, order_id=order_id, buy_amount=buy_amount, new_price=ema)
                update_order_id(new_order_info)
            
            else:
                # should be FILLED in that case
                if stop_trigger_toggle:
                    place_stop_triggers(client,  symbol_pair, last_close_price)
                    
                break

        old_remain_long_buy[symbol_pair] = remain